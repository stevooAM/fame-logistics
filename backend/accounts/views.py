"""
Accounts ViewSets — Invoice and Payment REST API.

Endpoints:
  GET  /api/accounts/invoices/           paginated, filterable list
  GET  /api/accounts/invoices/{id}/      full detail with nested payments
  POST /api/accounts/invoices/generate/  create invoice for approved job
  GET  /api/accounts/payments/           paginated list (filterable by invoice/customer)
  POST /api/accounts/payments/           record a payment, recalculate invoice status
  GET  /api/accounts/payments/{id}/      single payment detail

Permissions:
  Finance + Admin: read and write all
  Operations:      read-only on invoices; no payment creation, no invoice generation
  Other roles:     403

Audit:
  All write operations create AuditLog entries via AuditLogMixin._log_action.
"""

import csv
import io
import logging
from calendar import monthrange
from datetime import date as date_type
from datetime import timedelta
from decimal import Decimal

import openpyxl
from django.db.models import Sum
from django.db.models.functions import TruncMonth, TruncQuarter
from django.http import HttpResponse
from django.db import transaction
from django.utils import timezone
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from approvals.models import ApprovalQueue
from core.audit import AuditLogMixin
from core.permissions import IsAdminOrFinance, IsAnyRole
from jobs.models import Job

from .models import Invoice, Payment
from .serializers import (
    CustomerBalanceSerializer,
    GenerateInvoiceSerializer,
    InvoiceListSerializer,
    InvoiceSerializer,
    PaymentSerializer,
    PeriodSummaryResponseSerializer,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Pagination
# ---------------------------------------------------------------------------


class AccountsPagination(PageNumberPagination):
    """Server-side pagination for accounts lists — 20 records per page."""

    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 100


# ---------------------------------------------------------------------------
# Allowed ordering fields
# ---------------------------------------------------------------------------

_INVOICE_ALLOWED_ORDERING = {
    "invoice_number",
    "-invoice_number",
    "created_at",
    "-created_at",
    "status",
    "-status",
    "issue_date",
    "-issue_date",
    "due_date",
    "-due_date",
    "amount",
    "-amount",
    "customer__company_name",
    "-customer__company_name",
}


# ---------------------------------------------------------------------------
# InvoiceViewSet
# ---------------------------------------------------------------------------


class InvoiceViewSet(AuditLogMixin, ModelViewSet):
    """
    Invoice CRUD with an explicit generate action.

    list:      paginated, filterable, sortable
    retrieve:  single invoice with nested customer/job/payments + computed balance
    generate:  POST /invoices/generate/ — enforces "job must be approved" rule
    destroy:   405 — use CANCELLED status instead
    """

    pagination_class = AccountsPagination

    def get_permissions(self):
        """
        Permission matrix:
          list / retrieve / reporting → IsAnyRole  (Finance, Admin, Operations can read)
          all writes                  → IsAdminOrFinance  (Finance + Admin only)
        """
        read_only_actions = {
            "list",
            "retrieve",
            "balances",
            "balance_detail",
            "summary",
            "export_invoices",
            "export_balances",
        }
        if self.action in read_only_actions:
            return [IsAnyRole()]
        return [IsAdminOrFinance()]

    def get_serializer_class(self):
        """Lightweight ListSerializer for list; full InvoiceSerializer elsewhere."""
        if self.action == "list":
            return InvoiceListSerializer
        return InvoiceSerializer

    def get_queryset(self):
        """
        Build and filter the invoice queryset.

        Supports:
          search        OR across invoice_number, customer company name
          status        exact match on Invoice status
          customer_id   exact FK match
          job_id        exact FK match
          date_from     issue_date >= date
          date_to       issue_date <= date  (inclusive)
          ordering      allowed column (with optional - prefix)
        """
        qs = (
            Invoice.objects.select_related(
                "customer", "job", "currency", "created_by"
            )
            .prefetch_related("payments")
            .order_by("-created_at")
        )

        params = self.request.query_params

        # ---- Search --------------------------------------------------------
        search = params.get("search", "").strip()
        if search:
            from django.db.models import Q

            qs = qs.filter(
                Q(invoice_number__icontains=search)
                | Q(customer__company_name__icontains=search)
            )

        # ---- Exact filters -------------------------------------------------
        inv_status = params.get("status", "").strip()
        if inv_status:
            qs = qs.filter(status=inv_status)

        customer_id = params.get("customer_id", "").strip()
        if customer_id:
            try:
                qs = qs.filter(customer_id=int(customer_id))
            except (ValueError, TypeError):
                pass

        job_id = params.get("job_id", "").strip()
        if job_id:
            try:
                qs = qs.filter(job_id=int(job_id))
            except (ValueError, TypeError):
                pass

        # ---- Date range filters (on issue_date) ----------------------------
        date_from = params.get("date_from", "").strip()
        if date_from:
            try:
                qs = qs.filter(issue_date__gte=date_from)
            except (ValueError, TypeError):
                pass

        date_to = params.get("date_to", "").strip()
        if date_to:
            try:
                qs = qs.filter(issue_date__lte=date_to)
            except (ValueError, TypeError):
                pass

        # ---- Ordering ------------------------------------------------------
        ordering = params.get("ordering", "").strip()
        if ordering in _INVOICE_ALLOWED_ORDERING:
            qs = qs.order_by(ordering)

        return qs

    def destroy(self, request, *args, **kwargs):
        """Return 405 — invoices must use CANCELLED status instead of deletion."""
        return Response(
            {
                "detail": (
                    "Invoice deletion is not permitted. "
                    "Use status CANCELLED instead."
                )
            },
            status=status.HTTP_405_METHOD_NOT_ALLOWED,
        )

    def perform_create(self, serializer):
        """
        Save invoice with created_by set to the current user.

        Bypasses AuditLogMixin.perform_create to avoid double-logging —
        calls serializer.save() directly, then _log_action once.
        """
        serializer.save(created_by=self.request.user)
        self._log_action(self.request, "CREATE", serializer.instance)

    # -----------------------------------------------------------------------
    # Custom action: generate invoice for an approved job
    # -----------------------------------------------------------------------

    @action(detail=False, methods=["post"], url_path="generate")
    def generate(self, request):
        """
        POST /api/accounts/invoices/generate/

        Creates an Invoice in DRAFT status for a job that has at least one
        ApprovalQueue entry with status APPROVED.

        Request body:
          {
            "job_id":     <int>,
            "amount":     <decimal>,
            "currency_id": <int | null>,  (optional)
            "issue_date": <date | null>,   (optional — defaults to today)
            "due_date":   <date>,
            "notes":      <str>            (optional)
          }

        Returns 201 with the created Invoice (InvoiceSerializer).
        Returns 400 with {"detail": "Job has no approved approval entry"} if
        the job has never been approved.
        Returns 404 if the job_id does not exist.
        """
        serializer = GenerateInvoiceSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        data = serializer.validated_data
        job_id = data["job_id"]

        # --- Resolve job ---
        try:
            job = Job.objects.get(pk=job_id)
        except Job.DoesNotExist:
            return Response(
                {"detail": "Job not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        # --- Enforce approval gate ---
        has_approved = ApprovalQueue.objects.filter(
            job_id=job_id, status=ApprovalQueue.APPROVED
        ).exists()
        if not has_approved:
            return Response(
                {"detail": "Job has no approved approval entry"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # --- Create invoice atomically ---
        issue_date = data.get("issue_date") or timezone.now().date()

        with transaction.atomic():
            invoice = Invoice.objects.create(
                job=job,
                customer=job.customer,
                amount=data["amount"],
                currency_id=data.get("currency_id"),
                status=Invoice.DRAFT,
                issue_date=issue_date,
                due_date=data["due_date"],
                notes=data.get("notes", ""),
                created_by=request.user,
            )

        self._log_action(request, "CREATE", invoice)

        output_serializer = InvoiceSerializer(invoice, context={"request": request})
        return Response(output_serializer.data, status=status.HTTP_201_CREATED)


    # -----------------------------------------------------------------------
    # ACC-03: Customer outstanding-balances actions
    # -----------------------------------------------------------------------

    def _build_balance_rows(self, request):
        """
        Shared helper used by balances() and export_balances().

        Returns list[dict] with keys: customer_id, customer_name,
        invoiced_total, paid_total, balance, invoice_count, currency_code.

        Excludes CANCELLED invoices.
        Avoids N+1 via prefetch_related("invoices", "invoices__payments").
        """
        from customers.models import Customer

        base_invoices = Invoice.objects.exclude(status=Invoice.CANCELLED)

        customer_qs = (
            Customer.objects.filter(is_active=True, invoices__in=base_invoices)
            .select_related("currency_preference")
            .distinct()
        )

        # Optional filters
        customer_id_param = request.query_params.get("customer_id", "").strip()
        if customer_id_param:
            try:
                customer_qs = customer_qs.filter(pk=int(customer_id_param))
            except (TypeError, ValueError):
                pass

        search = request.query_params.get("search", "").strip()
        if search:
            customer_qs = customer_qs.filter(company_name__icontains=search)

        include_zero = request.query_params.get("include_zero", "").lower() == "true"
        zero = Decimal("0")
        rows = []

        customer_qs = customer_qs.prefetch_related("invoices", "invoices__payments")
        for c in customer_qs:
            invs = [i for i in c.invoices.all() if i.status != Invoice.CANCELLED]
            invoiced = sum((i.amount for i in invs), zero)
            paid = sum(
                (p.amount for i in invs for p in i.payments.all()),
                zero,
            )
            balance = invoiced - paid
            if not include_zero and balance == zero:
                continue
            pref = getattr(c, "currency_preference", None)
            currency_code = getattr(pref, "code", None) or "GHS"
            rows.append(
                {
                    "customer_id": c.id,
                    "customer_name": c.company_name,
                    "invoiced_total": invoiced,
                    "paid_total": paid,
                    "balance": balance,
                    "invoice_count": len(invs),
                    "currency_code": currency_code,
                }
            )

        ordering = request.query_params.get("ordering", "-balance").strip()
        reverse = ordering.startswith("-")
        key = ordering.lstrip("-")
        allowed_sort_keys = {
            "balance",
            "invoiced_total",
            "paid_total",
            "customer_name",
            "invoice_count",
        }
        if key not in allowed_sort_keys:
            key, reverse = "balance", True
        rows.sort(key=lambda r: (r[key], r["customer_name"]), reverse=reverse)
        return rows

    @action(detail=False, methods=["get"], url_path="balances")
    def balances(self, request):
        """
        GET /api/accounts/balances/

        One row per customer with a non-zero outstanding balance.
        Pass include_zero=true to include zero-balance customers.
        Sorted by balance DESC by default (pass ordering= to change).
        Paginated at 50/page (AccountsPagination page_size override).
        Each row includes currency_code (defaults to GHS).
        CANCELLED invoices excluded from all aggregations.
        """
        rows = self._build_balance_rows(request)

        paginator = AccountsPagination()
        paginator.page_size = 50
        page = paginator.paginate_queryset(rows, request, view=self)
        if page is not None:
            return paginator.get_paginated_response(
                CustomerBalanceSerializer(page, many=True).data
            )
        return Response(CustomerBalanceSerializer(rows, many=True).data)

    @action(
        detail=False,
        methods=["get"],
        url_path=r"balances/(?P<customer_id>[0-9]+)",
    )
    def balance_detail(self, request, customer_id=None):
        """
        GET /api/accounts/balances/{customer_id}/

        Aggregate + invoice list for one customer.
        Customer dict includes id, company_name, email, phone_number, tin.
        Invoices are serialized via InvoiceListSerializer (paid_total + balance fields).
        """
        from customers.models import Customer

        try:
            customer = Customer.objects.get(pk=customer_id)
        except Customer.DoesNotExist:
            return Response(
                {"detail": "Customer not found."},
                status=status.HTTP_404_NOT_FOUND,
            )

        agg = Invoice.objects.outstanding_for_customer(int(customer_id))
        invoices = (
            Invoice.objects.filter(customer_id=customer_id)
            .exclude(status=Invoice.CANCELLED)
            .select_related("job", "currency")
            .prefetch_related("payments")
            .order_by("-issue_date")
        )
        return Response(
            {
                "customer": {
                    "id": customer.id,
                    "company_name": customer.company_name,
                    # customer.phone is the actual model field (see customers/models.py).
                    # Exposed as phone_number in the API per the locked contract.
                    "email": customer.email,
                    "phone_number": customer.phone,
                    "tin": customer.tin,
                },
                "invoiced_total": str(agg["invoiced"]),
                "paid_total": str(agg["paid"]),
                "balance": str(agg["balance"]),
                "invoice_count": invoices.count(),
                "invoices": InvoiceListSerializer(invoices, many=True).data,
            }
        )

    # -----------------------------------------------------------------------
    # ACC-04: Period summary action
    # -----------------------------------------------------------------------

    @action(detail=False, methods=["get"], url_path="summary")
    def summary(self, request):
        """
        GET /api/accounts/summary/?period=month|quarter[&date_from=&date_to=]

        Returns WRAPPED response per locked API contract:
          {
            "period": "month" | "quarter",
            "date_from": "YYYY-MM-DD" | "",
            "date_to": "YYYY-MM-DD" | "",
            "rows": [{period_label, period_start, period_end, invoiced, paid, outstanding}],
            "totals": {invoiced, paid, outstanding}
          }

        Up to 12 month rows or 8 quarter rows in chronological ascending order.
        Query params: date_from / date_to (NOT start_date / end_date).
        CANCELLED invoices excluded.
        """
        period = request.query_params.get("period", "month").lower()
        if period not in ("month", "quarter"):
            period = "month"
        trunc_fn = TruncMonth if period == "month" else TruncQuarter
        limit = 12 if period == "month" else 8

        invoice_qs = Invoice.objects.exclude(status=Invoice.CANCELLED)
        payment_qs = Payment.objects.exclude(invoice__status=Invoice.CANCELLED)

        df = request.query_params.get("date_from", "").strip()
        dt = request.query_params.get("date_to", "").strip()
        if df:
            invoice_qs = invoice_qs.filter(issue_date__gte=df)
            payment_qs = payment_qs.filter(payment_date__gte=df)
        if dt:
            invoice_qs = invoice_qs.filter(issue_date__lte=dt)
            payment_qs = payment_qs.filter(payment_date__lte=dt)

        invoiced_by_period = (
            invoice_qs.annotate(period=trunc_fn("issue_date"))
            .values("period")
            .annotate(total=Sum("amount"))
        )
        paid_by_period = (
            payment_qs.annotate(period=trunc_fn("payment_date"))
            .values("period")
            .annotate(total=Sum("amount"))
        )

        bucket = {}
        for row in invoiced_by_period:
            bucket.setdefault(
                row["period"], {"invoiced": Decimal("0"), "paid": Decimal("0")}
            )
            bucket[row["period"]]["invoiced"] = row["total"] or Decimal("0")
        for row in paid_by_period:
            bucket.setdefault(
                row["period"], {"invoiced": Decimal("0"), "paid": Decimal("0")}
            )
            bucket[row["period"]]["paid"] = row["total"] or Decimal("0")

        rows = []
        for start, agg in sorted(bucket.items(), reverse=True)[:limit]:
            if period == "month":
                label = start.strftime("%b %Y")
                end = start.replace(day=monthrange(start.year, start.month)[1])
            else:  # quarter
                q = (start.month - 1) // 3 + 1
                label = f"Q{q} {start.year}"
                end_month = start.month + 2
                end = start.replace(
                    month=end_month,
                    day=monthrange(start.year, end_month)[1],
                )
            rows.append(
                {
                    "period_label": label,
                    "period_start": start,
                    "period_end": end,
                    "invoiced": agg["invoiced"],
                    "paid": agg["paid"],
                    "outstanding": agg["invoiced"] - agg["paid"],
                }
            )
        rows.reverse()  # chronological ascending in response

        zero = Decimal("0")
        totals = {
            "invoiced": sum((r["invoiced"] for r in rows), zero),
            "paid": sum((r["paid"] for r in rows), zero),
            "outstanding": sum((r["outstanding"] for r in rows), zero),
        }

        payload = {
            "period": period,
            "date_from": df,
            "date_to": dt,
            "rows": rows,
            "totals": totals,
        }
        return Response(PeriodSummaryResponseSerializer(payload).data)

    # -----------------------------------------------------------------------
    # ACC-05: Excel/CSV export actions
    # -----------------------------------------------------------------------

    @action(detail=False, methods=["get"], url_path="export")
    def export_invoices(self, request):
        """
        GET /api/accounts/invoices/export/?format=xlsx|csv

        Exports the filtered invoice list. Reuses get_queryset() so all
        filter params (search, status, customer_id, job_id, date_from,
        date_to, ordering) are honoured. No pagination — returns all rows.
        No audit log (read-only operation — consistent with Phase 4 decision).
        """
        fmt = request.query_params.get("format", "xlsx").lower()
        if fmt not in ("xlsx", "csv"):
            fmt = "xlsx"
        qs = self.get_queryset()

        headers = [
            "Invoice Number",
            "Job Number",
            "Customer",
            "Amount",
            "Currency",
            "Paid",
            "Balance",
            "Status",
            "Issue Date",
            "Due Date",
            "Created At",
        ]

        def _row(inv):
            return [
                inv.invoice_number,
                inv.job.job_number if inv.job_id else "",
                inv.customer.company_name if inv.customer_id else "",
                float(inv.amount) if inv.amount is not None else 0,
                getattr(inv.currency, "code", "") if inv.currency_id else "",
                float(inv.paid_total()),
                float(inv.balance()),
                inv.status,
                inv.issue_date.strftime("%Y-%m-%d") if inv.issue_date else "",
                inv.due_date.strftime("%Y-%m-%d") if inv.due_date else "",
                (
                    inv.created_at.strftime("%Y-%m-%d %H:%M:%S")
                    if inv.created_at
                    else ""
                ),
            ]

        today = date_type.today().strftime("%Y%m%d")

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Invoices"
            ws.append(headers)
            bold = openpyxl.styles.Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
            widths = [18, 18, 30, 14, 10, 14, 14, 12, 14, 14, 22]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            for inv in qs:
                ws.append(_row(inv))
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(
                buf.getvalue(),
                content_type=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
            )
            resp["Content-Disposition"] = (
                f'attachment; filename="invoices_export_{today}.xlsx"'
            )
            return resp

        # CSV
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for inv in qs:
            w.writerow(_row(inv))
        resp = HttpResponse(buf.getvalue(), content_type="text/csv")
        resp["Content-Disposition"] = (
            f'attachment; filename="invoices_export_{today}.csv"'
        )
        return resp

    @action(detail=False, methods=["get"], url_path=r"balances/export")
    def export_balances(self, request):
        """
        GET /api/accounts/balances/export/?format=xlsx|csv

        Exports all customer balance rows (same filters as /balances/ but
        no pagination). Includes Currency column (currency_code per row).
        No audit log (read-only operation — consistent with Phase 4 decision).
        """
        fmt = request.query_params.get("format", "xlsx").lower()
        if fmt not in ("xlsx", "csv"):
            fmt = "xlsx"
        rows = self._build_balance_rows(request)

        headers = [
            "Customer",
            "Currency",
            "Invoice Count",
            "Invoiced Total",
            "Paid Total",
            "Balance",
        ]
        today = date_type.today().strftime("%Y%m%d")

        if fmt == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Outstanding Balances"
            ws.append(headers)
            bold = openpyxl.styles.Font(bold=True)
            for cell in ws[1]:
                cell.font = bold
            for i, w in enumerate([35, 10, 14, 16, 16, 16], start=1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
            for r in rows:
                ws.append(
                    [
                        r["customer_name"],
                        r.get("currency_code", "GHS"),
                        r["invoice_count"],
                        float(r["invoiced_total"]),
                        float(r["paid_total"]),
                        float(r["balance"]),
                    ]
                )
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(
                buf.getvalue(),
                content_type=(
                    "application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"
                ),
            )
            resp["Content-Disposition"] = (
                f'attachment; filename="balances_export_{today}.xlsx"'
            )
            return resp

        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(headers)
        for r in rows:
            w.writerow(
                [
                    r["customer_name"],
                    r.get("currency_code", "GHS"),
                    r["invoice_count"],
                    f"{r['invoiced_total']:.2f}",
                    f"{r['paid_total']:.2f}",
                    f"{r['balance']:.2f}",
                ]
            )
        resp = HttpResponse(buf.getvalue(), content_type="text/csv")
        resp["Content-Disposition"] = (
            f'attachment; filename="balances_export_{today}.csv"'
        )
        return resp


# ---------------------------------------------------------------------------
# PaymentViewSet
# ---------------------------------------------------------------------------


class PaymentViewSet(AuditLogMixin, ModelViewSet):
    """
    Payment recording and retrieval.

    list:    paginated list (filterable by invoice_id, customer_id)
    create:  record payment → recalculate Invoice.status atomically
    retrieve: single payment detail

    PUT / PATCH / DELETE are disabled.
    """

    pagination_class = AccountsPagination
    # Restrict to safe HTTP methods — payments are immutable once recorded
    http_method_names = ["get", "post", "head", "options"]

    def get_permissions(self):
        """
        Permission matrix:
          list / retrieve → IsAnyRole        (all staff can read)
          create          → IsAdminOrFinance  (Finance + Admin only)
        """
        read_only_actions = {"list", "retrieve"}
        if self.action in read_only_actions:
            return [IsAnyRole()]
        return [IsAdminOrFinance()]

    def get_serializer_class(self):
        return PaymentSerializer

    def get_queryset(self):
        """
        Build and filter the payment queryset.

        Supports:
          invoice_id    exact FK match
          customer_id   filter through invoice__customer_id
        """
        qs = Payment.objects.select_related(
            "invoice", "invoice__customer", "recorded_by"
        ).order_by("-payment_date")

        params = self.request.query_params

        invoice_id = params.get("invoice_id", "").strip()
        if invoice_id:
            try:
                qs = qs.filter(invoice_id=int(invoice_id))
            except (ValueError, TypeError):
                pass

        customer_id = params.get("customer_id", "").strip()
        if customer_id:
            try:
                qs = qs.filter(invoice__customer_id=int(customer_id))
            except (ValueError, TypeError):
                pass

        return qs

    def perform_create(self, serializer):
        """
        Record a payment and atomically recalculate the parent invoice status.

        Status recalculation rules:
          balance == 0              → PAID
          0 < paid < invoiced       → PARTIAL
          otherwise (no payments yet, or data inconsistency) → SENT

        Bypasses AuditLogMixin.perform_create to avoid double-logging.
        """
        with transaction.atomic():
            payment = serializer.save(recorded_by=self.request.user)

            # Recalculate parent invoice status
            invoice = payment.invoice
            # Re-fetch inside the transaction to get accurate aggregates
            invoice.refresh_from_db()
            balance = invoice.balance()
            paid = invoice.paid_total()
            invoiced = invoice.invoiced_total()

            if balance == 0:
                new_status = Invoice.PAID
            elif 0 < paid < invoiced:
                new_status = Invoice.PARTIAL
            else:
                new_status = Invoice.SENT

            invoice.status = new_status
            invoice.save(update_fields=["status", "updated_at"])

        self._log_action(self.request, "CREATE", payment)
