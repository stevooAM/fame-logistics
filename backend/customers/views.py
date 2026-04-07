"""
Customer ViewSet — CRUD with pagination, filtering, soft-delete, and TIN duplicate check.

Permission: IsAdminOrOperations (Admin + Operations roles)
Audit:      AuditLogMixin auto-logs CREATE/UPDATE; destroy() calls _log_action manually
Pagination: 20 per page (page_size_query_param allows override up to 100)
"""

import csv
import io
import logging
from datetime import date

import openpyxl
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from core.audit import AuditLogMixin
from core.permissions import IsAdminOrOperations

from .models import Customer
from .serializers import CustomerListSerializer, CustomerSerializer

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Pagination
# ---------------------------------------------------------------------------


class CustomerPagination(PageNumberPagination):
    """Server-side pagination for customer list — 20 records per page by default."""

    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 100


# ---------------------------------------------------------------------------
# ViewSet
# ---------------------------------------------------------------------------

#: Allowed ordering fields — prefix with "-" for descending
_ALLOWED_ORDERING = {
    "company_name",
    "-company_name",
    "tin",
    "-tin",
    "created_at",
    "-created_at",
}


class CustomerViewSet(AuditLogMixin, ModelViewSet):
    """
    Full CRUD ViewSet for Customer master records.

    list:       paginated, filtered, sortable
    retrieve:   single customer with nested FK detail
    create:     creates and audit-logs
    update:     full update (PUT) and audit-logs
    partial_update: PATCH and audit-logs
    destroy:    soft-delete (is_active=False), returns 200 with updated object
    check-tin:  custom action — duplicate TIN check
    """

    permission_classes = [IsAdminOrOperations]
    pagination_class = CustomerPagination

    def get_queryset(self):
        """
        Build and filter the base queryset.

        By default only active customers are returned. Passing
        ``?include_inactive=true`` includes soft-deleted records (for admin
        audit / recovery UIs).
        """
        qs = Customer.objects.select_related(
            "preferred_port", "currency_preference"
        ).all()

        # Active-only gate
        include_inactive = (
            self.request.query_params.get("include_inactive", "").lower() == "true"
        )
        if not include_inactive:
            qs = qs.filter(is_active=True)

        # ---- Search (OR across key columns) --------------------------------
        search = self.request.query_params.get("search", "").strip()
        if search:
            qs = qs.filter(
                Q(company_name__icontains=search)
                | Q(tin__icontains=search)
                | Q(contact_person__icontains=search)
                | Q(email__icontains=search)
            )

        # ---- Column filters -------------------------------------------------
        company_name = self.request.query_params.get("company_name", "").strip()
        if company_name:
            qs = qs.filter(company_name__icontains=company_name)

        tin = self.request.query_params.get("tin", "").strip()
        if tin:
            qs = qs.filter(tin__icontains=tin)

        business_type = self.request.query_params.get("business_type", "").strip()
        if business_type:
            qs = qs.filter(business_type__icontains=business_type)

        customer_type = self.request.query_params.get("customer_type", "").strip()
        if customer_type:
            qs = qs.filter(customer_type=customer_type)

        credit_terms = self.request.query_params.get("credit_terms", "").strip()
        if credit_terms:
            qs = qs.filter(credit_terms__icontains=credit_terms)

        # ---- Ordering -------------------------------------------------------
        ordering = self.request.query_params.get("ordering", "company_name").strip()
        if ordering in _ALLOWED_ORDERING:
            qs = qs.order_by(ordering)
        else:
            qs = qs.order_by("company_name")

        return qs

    def get_serializer_class(self):
        """Use lighter ListSerializer for list action; full serializer elsewhere."""
        if self.action == "list":
            return CustomerListSerializer
        return CustomerSerializer

    # -----------------------------------------------------------------------
    # Soft-delete override
    # -----------------------------------------------------------------------

    def destroy(self, request, *args, **kwargs):
        """
        Soft-delete: set is_active=False, audit-log DELETE, return 200.

        Matches the pattern established in setup/views.py (LookupViewSetMixin)
        so the frontend can update list state without a re-fetch.
        """
        instance = self.get_object()
        instance.is_active = False
        instance.save(update_fields=["is_active", "updated_at"])
        self._log_action(request, "DELETE", instance)
        serializer = self.get_serializer(instance)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # -----------------------------------------------------------------------
    # Custom actions
    # -----------------------------------------------------------------------

    @action(detail=False, methods=["get"], url_path="check-tin")
    def check_tin(self, request):
        """
        GET /api/customers/check-tin/?tin=XXX[&exclude_id=YY]

        Returns whether a TIN already exists in the system so the frontend
        can warn the user before they submit the create/edit form.

        Searches ALL customers (including inactive) to prevent re-use of
        previously deactivated customer TINs.

        If ``exclude_id`` is provided the customer with that PK is excluded
        from the search (edit-mode — don't flag the customer's own TIN as duplicate).

        Response:
            {
                "duplicate": true | false,
                "existing_customer": {"id": N, "company_name": "..."} | null
            }
        """
        tin = request.query_params.get("tin", "").strip()
        exclude_id = request.query_params.get("exclude_id", "").strip()

        if not tin:
            return Response(
                {"duplicate": False, "existing_customer": None},
                status=status.HTTP_200_OK,
            )

        qs = Customer.objects.filter(tin=tin)

        if exclude_id:
            try:
                qs = qs.exclude(pk=int(exclude_id))
            except (ValueError, TypeError):
                pass  # Invalid exclude_id — treat as not provided

        existing = qs.values("id", "company_name").first()

        if existing:
            return Response(
                {
                    "duplicate": True,
                    "existing_customer": {
                        "id": existing["id"],
                        "company_name": existing["company_name"],
                    },
                },
                status=status.HTTP_200_OK,
            )

        return Response(
            {"duplicate": False, "existing_customer": None},
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="export")
    def export_customers(self, request):
        """
        GET /api/customers/export/[?format=xlsx|csv]

        Exports all customer records matching the current filter params.
        Pagination is NOT applied — all matching rows are returned.

        Supported formats:
          - xlsx (default): Excel workbook with bold header row
          - csv: Plain CSV

        Columns: Company Name, TIN, Customer Type, Contact Person, Phone,
                 Email, Address, Business Type, Credit Terms, Notes,
                 Active Status, Created At

        Permission: IsAdminOrOperations (inherited from ViewSet)
        Audit: Not logged — read-only export operation
        """
        export_format = request.query_params.get("format", "xlsx").lower()
        if export_format not in ("xlsx", "csv"):
            export_format = "xlsx"

        # Reuse full filtered queryset (no pagination)
        qs = self.get_queryset()

        # Extract values for all export columns in one DB hit
        rows = qs.values(
            "company_name",
            "tin",
            "customer_type",
            "contact_person",
            "phone",
            "email",
            "address",
            "business_type",
            "credit_terms",
            "notes",
            "is_active",
            "created_at",
        )

        headers = [
            "Company Name",
            "TIN",
            "Customer Type",
            "Contact Person",
            "Phone",
            "Email",
            "Address",
            "Business Type",
            "Credit Terms",
            "Notes",
            "Active Status",
            "Created At",
        ]

        today = date.today().strftime("%Y%m%d")

        # -----------------------------------------------------------------------
        # XLSX
        # -----------------------------------------------------------------------
        if export_format == "xlsx":
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Customers"

            # Bold header row
            bold_font = openpyxl.styles.Font(bold=True)
            ws.append(headers)
            for cell in ws[1]:
                cell.font = bold_font

            # Column widths
            col_widths = [35, 20, 18, 25, 18, 30, 40, 25, 20, 40, 12, 22]
            for idx, width in enumerate(col_widths, start=1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(idx)
                ].width = width

            for row in rows:
                created_at = row["created_at"]
                if created_at is not None:
                    # Strip timezone for Excel compatibility
                    created_at = created_at.strftime("%Y-%m-%d %H:%M:%S")
                ws.append(
                    [
                        row["company_name"],
                        row["tin"],
                        row["customer_type"],
                        row["contact_person"],
                        row["phone"],
                        row["email"],
                        row["address"],
                        row["business_type"],
                        row["credit_terms"],
                        row["notes"],
                        "Yes" if row["is_active"] else "No",
                        created_at,
                    ]
                )

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            response = HttpResponse(
                buffer.getvalue(),
                content_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                ),
            )
            response["Content-Disposition"] = (
                f'attachment; filename="customers_export_{today}.xlsx"'
            )
            return response

        # -----------------------------------------------------------------------
        # CSV
        # -----------------------------------------------------------------------
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(headers)

        for row in rows:
            created_at = row["created_at"]
            if created_at is not None:
                created_at = created_at.strftime("%Y-%m-%d %H:%M:%S")
            writer.writerow(
                [
                    row["company_name"],
                    row["tin"],
                    row["customer_type"],
                    row["contact_person"],
                    row["phone"],
                    row["email"],
                    row["address"],
                    row["business_type"],
                    row["credit_terms"],
                    row["notes"],
                    "Yes" if row["is_active"] else "No",
                    created_at,
                ]
            )

        response = HttpResponse(buffer.getvalue(), content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="customers_export_{today}.csv"'
        )
        return response
