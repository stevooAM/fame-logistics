"""
Report API views for Fame FMS.

Endpoints:
  GET /api/reports/customer-activity/   RPT-01
  GET /api/reports/job-status/           RPT-02
  GET /api/reports/revenue/              RPT-03

All endpoints accept:
  date_from=YYYY-MM-DD  (default: first day of current month)
  date_to=YYYY-MM-DD    (default: last day of current month)

Revenue endpoint also accepts:
  currency_code=GHS|USD|...  (default: all currencies combined)

Customer activity endpoint also accepts:
  customer_id=<int>  (default: all customers)

Permissions: IsAnyRole (Admin, Finance, Operations)
"""

from datetime import date, datetime
from calendar import monthrange

from rest_framework.response import Response
from rest_framework.views import APIView

from core.permissions import IsAnyRole

from .queries import customer_activity_query, job_status_query, revenue_query


def _parse_dates(request):
    """
    Parse date_from and date_to from query params.
    Defaults to first/last day of current month when params absent.
    Returns (date_from, date_to) on success or (None, None, error_response) on failure.
    """
    today = date.today()
    first_of_month = today.replace(day=1)
    _, last_day = monthrange(today.year, today.month)
    last_of_month = today.replace(day=last_day)

    raw_from = request.query_params.get("date_from")
    raw_to = request.query_params.get("date_to")

    try:
        date_from = datetime.strptime(raw_from, "%Y-%m-%d").date() if raw_from else first_of_month
        date_to = datetime.strptime(raw_to, "%Y-%m-%d").date() if raw_to else last_of_month
    except ValueError:
        return None, None, Response(
            {"error": "Invalid date format. Use YYYY-MM-DD."},
            status=400,
        )

    if date_from > date_to:
        return None, None, Response(
            {"error": "date_from must be on or before date_to."},
            status=400,
        )

    return date_from, date_to


class CustomerActivityView(APIView):
    permission_classes = [IsAnyRole]

    def get(self, request):
        result = _parse_dates(request)
        if len(result) == 3:
            return result[2]
        date_from, date_to = result

        customer_id = request.query_params.get("customer_id")
        if customer_id:
            try:
                customer_id = int(customer_id)
            except (ValueError, TypeError):
                return Response({"error": "customer_id must be an integer."}, status=400)

        rows = customer_activity_query(date_from, date_to, customer_id=customer_id)
        return Response({
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "rows": rows,
            "count": len(rows),
        })


class JobStatusView(APIView):
    permission_classes = [IsAnyRole]

    def get(self, request):
        result = _parse_dates(request)
        if len(result) == 3:
            return result[2]
        date_from, date_to = result

        rows = job_status_query(date_from, date_to)
        return Response({
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "rows": rows,
        })


class RevenueView(APIView):
    permission_classes = [IsAnyRole]

    def get(self, request):
        result = _parse_dates(request)
        if len(result) == 3:
            return result[2]
        date_from, date_to = result

        currency_code = request.query_params.get("currency_code") or None

        data = revenue_query(date_from, date_to, currency_code=currency_code)
        return Response({
            "date_from": date_from.isoformat(),
            "date_to": date_to.isoformat(),
            "currency_code": currency_code,
            **data,
        })


# ---------------------------------------------------------------------------
# Export views — PDF (WeasyPrint) + Excel (openpyxl)
# ---------------------------------------------------------------------------

import io  # noqa: E402

import openpyxl  # noqa: E402
from django.http import HttpResponse  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

from .pdf_utils import generate_report_pdf  # noqa: E402


class CustomerActivityExportView(APIView):
    permission_classes = [IsAnyRole]

    def get(self, request):
        result = _parse_dates(request)
        if len(result) == 3:
            return result[2]
        date_from, date_to = result

        customer_id = request.query_params.get("customer_id")
        if customer_id:
            try:
                customer_id = int(customer_id)
            except (ValueError, TypeError):
                return Response({"error": "customer_id must be an integer."}, status=400)

        fmt = request.query_params.get("format", "xlsx").lower()
        rows = customer_activity_query(date_from, date_to, customer_id=customer_id)

        if fmt == "pdf":
            pdf_bytes = generate_report_pdf("customer-activity", rows, date_from, date_to)
            resp = HttpResponse(pdf_bytes, content_type="application/pdf")
            resp["Content-Disposition"] = (
                f'attachment; filename="customer-activity-{date_from}-{date_to}.pdf"'
            )
            return resp

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Activity"
        headers = [
            "Customer", "Total Jobs", "Total Value",
            "Draft", "Pending", "In Progress", "Customs",
            "Delivered", "Closed", "Cancelled",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([
                r["customer_name"],
                r["total_jobs"],
                float(r["total_value"]),
                r.get("draft", 0),
                r.get("pending", 0),
                r.get("in_progress", 0),
                r.get("customs", 0),
                r.get("delivered", 0),
                r.get("closed", 0),
                r.get("cancelled", 0),
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="customer-activity-{date_from}-{date_to}.xlsx"'
        )
        return resp


class JobStatusExportView(APIView):
    permission_classes = [IsAnyRole]

    def get(self, request):
        result = _parse_dates(request)
        if len(result) == 3:
            return result[2]
        date_from, date_to = result

        fmt = request.query_params.get("format", "xlsx").lower()
        rows = job_status_query(date_from, date_to)

        if fmt == "pdf":
            pdf_bytes = generate_report_pdf("job-status", rows, date_from, date_to)
            resp = HttpResponse(pdf_bytes, content_type="application/pdf")
            resp["Content-Disposition"] = (
                f'attachment; filename="job-status-{date_from}-{date_to}.pdf"'
            )
            return resp

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Job Status"
        headers = ["Status", "Job Type", "Count", "Total Value"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for r in rows:
            ws.append([r["status_label"], r["job_type_label"], r["count"], float(r["total_value"])])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="job-status-{date_from}-{date_to}.xlsx"'
        )
        return resp


class RevenueExportView(APIView):
    permission_classes = [IsAnyRole]

    def get(self, request):
        result = _parse_dates(request)
        if len(result) == 3:
            return result[2]
        date_from, date_to = result

        currency_code = request.query_params.get("currency_code") or None
        fmt = request.query_params.get("format", "xlsx").lower()
        data = revenue_query(date_from, date_to, currency_code=currency_code)

        if fmt == "pdf":
            pdf_bytes = generate_report_pdf(
                "revenue", data, date_from, date_to, currency_code=currency_code
            )
            resp = HttpResponse(pdf_bytes, content_type="application/pdf")
            resp["Content-Disposition"] = (
                f'attachment; filename="revenue-{date_from}-{date_to}.pdf"'
            )
            return resp

        wb = openpyxl.Workbook()

        ws1 = wb.active
        ws1.title = "Period Summary"
        ws1.append(["Period", "Invoiced", "Paid", "Outstanding"])
        for cell in ws1[1]:
            cell.font = Font(bold=True)
        for r in data.get("period_rows", []):
            ws1.append([r["period_label"], float(r["invoiced"]), float(r["paid"]), float(r["outstanding"])])

        ws2 = wb.create_sheet("Customer Breakdown")
        ws2.append(["Customer", "Invoiced", "Paid", "Outstanding"])
        for cell in ws2[1]:
            cell.font = Font(bold=True)
        for r in data.get("customer_rows", []):
            ws2.append([r["customer_name"], float(r["invoiced"]), float(r["paid"]), float(r["outstanding"])])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        resp = HttpResponse(
            buf.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = (
            f'attachment; filename="revenue-{date_from}-{date_to}.xlsx"'
        )
        return resp
